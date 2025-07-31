"""Excel匯出工具"""
import os
import logging
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config.settings import Settings
from models.revenue_store import RevenueStore

class ExcelExporter:
    def __init__(self, settings: Settings):
        """初始化Excel匯出器
        
        Args:
            settings: 設定物件
        """
        self.settings = settings
        self.excel_config = settings.get_excel_config()
    
    def export_revenue_data(self, revenue_store: RevenueStore, output_path: str = None) -> bool:
        """匯出營收數據到Excel
        
        Args:
            revenue_store: 營收資料庫
            output_path: 輸出檔案路徑，如果為None則自動生成
            
        Returns:
            bool: 是否匯出成功
        """
        try:
            print("\n=== 匯出營收數據到Excel ===")
            
            # 生成輸出檔案路徑
            if output_path is None:
                output_path = self._generate_output_path()
            
            # 獲取營收數據
            revenue_data = revenue_store.get_all_revenue_data()
            
            if not revenue_data:
                print("沒有營收數據可以匯出")
                return False
            
            # 創建Excel檔案
            wb = Workbook()
            
            # 移除預設工作表
            if wb.active:
                wb.remove(wb.active)
            
            # 創建總營收工作表
            ws = wb.create_sheet(title="總營收")
            
            # 設定樣式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # 寫入表頭
            headers = self.excel_config.get('headers', [
                '公司', '年份季度', '總營收(百萬美元)', '原始數值', '原始單位', '原始幣別', '數據類型', '建立日期'
            ])
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 寫入數據
            for row_num, record in enumerate(revenue_data, 2):
                ws.cell(row=row_num, column=1).value = record.get('company')
                ws.cell(row=row_num, column=2).value = record.get('year_quarter')
                ws.cell(row=row_num, column=3).value = record.get('value')
                ws.cell(row=row_num, column=4).value = record.get('original_value', record.get('value'))
                ws.cell(row=row_num, column=5).value = record.get('original_unit', '')
                ws.cell(row=row_num, column=6).value = record.get('original_currency', 'USD')
                ws.cell(row=row_num, column=7).value = record.get('data_type', 'actual')
                ws.cell(row=row_num, column=8).value = record.get('created_at')
            
            # 調整列寬
            self._adjust_column_widths(ws)
            
            # 保存文件
            wb.save(output_path)
            print(f"營收數據已成功匯出到: {output_path}")
            
            # 顯示統計信息
            print(f"總營收記錄: {len(revenue_data)}")
            
            # 統計不同類型的記錄
            actual_count = sum(1 for record in revenue_data if record.get('data_type') == 'actual')
            fullyear_count = sum(1 for record in revenue_data if record.get('data_type') == 'actual_fullyear')
            q4_count = sum(1 for record in revenue_data if record.get('year_quarter', '').endswith('_Q4'))
            
            print(f"  - 季度記錄: {actual_count}")
            print(f"  - 年度記錄: {fullyear_count}")
            print(f"  - Q4記錄: {q4_count}")
            
            return True
            
        except Exception as e:
            logging.error(f"匯出Excel時發生錯誤: {e}")
            return False
    
    def _generate_output_path(self) -> str:
        """生成輸出檔案路徑
        
        Returns:
            str: 輸出檔案路徑
        """
        current_time = datetime.now()
        datetime_version = current_time.strftime("%m%d%H%M")
        filename_prefix = self.excel_config.get('filename_prefix', '競業營收數據')
        filename = f"{filename_prefix}_{datetime_version}.xlsx"
        return os.path.join(os.getcwd(), filename)
    
    def _adjust_column_widths(self, worksheet):
        """調整列寬
        
        Args:
            worksheet: 工作表物件
        """
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # 設定適當的列寬，最小10，最大50
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logging.warning(f"調整列寬時發生錯誤: {e}")