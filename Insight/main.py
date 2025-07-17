import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config.settings import settings
from utils.logger import setup_logger, get_logger
from utils.file_utils import (
    find_report_folders, find_pdf_files, extract_company_name,
    extract_year_and_quarter, create_output_directory, generate_excel_filename
)
from models.vector_store import EnhancedMongoDBVectorStore
from analyzers.rag_analyzer import RAGAnalyzer

# 設置日誌
setup_logger()
logger = get_logger(__name__)

def test_connections():
    """測試API和模型連接"""
    try:
        # 測試OpenAI API
        from openai import OpenAI
        client = OpenAI(api_key=settings.openai_api_key)
        test_response = client.chat.completions.create(
            model=settings.llm_model,
            messages=[{"role": "user", "content": "測試連接"}],
            max_tokens=10
        )
        logger.info(f"{settings.llm_model} API 連接成功")
        
        # 測試SentenceTransformer
        from sentence_transformers import SentenceTransformer
        embedding_model = SentenceTransformer(settings.embedding_model)
        test_embedding = embedding_model.encode("測試", convert_to_tensor=False)
        logger.info(f"SentenceTransformer ({settings.embedding_model}) 載入成功，維度: {len(test_embedding)}")
        
        return True
        
    except Exception as e:
        logger.error(f"連接測試失敗: {e}")
        return False

def check_processed_files(vector_store):
    """檢查哪些檔案已經處理過"""
    pipeline = [
        {
            "$group": {
                "_id": {
                    "company": "$metadata.company_name",
                    "file_name": "$metadata.file_name"
                }
            }
        }
    ]
    
    processed_files = list(vector_store.collection.aggregate(pipeline))
    processed_set = set()
    
    for item in processed_files:
        company = item['_id']['company']
        file_name = item['_id']['file_name']
        processed_set.add(f"{company}_{file_name}")
    
    return processed_set

def preprocess_all_companies(vector_store, force_reprocess=False):
    """前處理：將所有公司的財報資料處理並存入MongoDB"""
    logger.info("=== 處理所有公司財報資料 ===")
    
    # 檢查是否已有資料
    existing_docs_count = vector_store.collection.count_documents({})
    
    if existing_docs_count > 0 and not force_reprocess:
        logger.info(f"發現資料庫中已有 {existing_docs_count} 個文檔塊")
        user_choice = input("是否要重新處理所有檔案？(y/N): ").lower()
        
        if user_choice != 'y':
            logger.info("跳過前處理階段，使用現有資料")
            return True
        else:
            logger.info("清空向量資料庫，重新處理...")
            vector_store.clear_collection()
    elif force_reprocess:
        logger.info("強制重新處理，清空向量資料庫...")
        vector_store.clear_collection()
    
    # 尋找財報資料夾
    report_folders = find_report_folders(settings.base_directory)
    
    if not report_folders:
        logger.error("找不到任何財報資料夾")
        return False
    
    logger.info(f"找到 {len(report_folders)} 個財報資料夾")
    
    total_processed = 0
    total_failed = 0
    
    # 處理每個財報資料夾
    for folder in report_folders:
        company_name = extract_company_name(folder)
        logger.info(f"\n開始前處理 {company_name} 的財報...")
        
        # 處理PDF文件
        pdf_files = find_pdf_files(folder)
        logger.info(f"找到 {len(pdf_files)} 個 PDF 文件")
        
        for index, pdf_file in enumerate(pdf_files, 1):
            file_name = os.path.basename(pdf_file)
            year, quarter = extract_year_and_quarter(file_name)
            
            if not year or not quarter:
                logger.warning(f"無法提取年份季度信息，跳過: {file_name}")
                total_failed += 1
                continue
                
            logger.info(f"前處理第 {index}/{len(pdf_files)} 個檔案: {file_name}")
            
            # 智能處理流程
            max_attempts = 2  # 最多嘗試2次（文字提取 + OCR）
            current_attempt = 1
            success = False
            
            while current_attempt <= max_attempts and not success:
                try:
                    logger.info(f"嘗試第 {current_attempt} 次處理: {file_name}")
                    
                    # 使用智能PDF讀取
                    pdf_text, total_pages = vector_store.read_pdf_enhanced(pdf_file, company_name)
                    if not pdf_text:
                        logger.warning(f"無法讀取 {file_name}，跳過")
                        break
                    
                    # 添加到向量資料庫
                    use_ocr = vector_store.should_use_ocr_processing(file_name, company_name)
                    metadata = {
                        "file_name": file_name,
                        "company_name": company_name,
                        "year": year,
                        "quarter": f"{year}_{quarter}",
                        "total_pages": total_pages,
                        "processing_mode": "ocr_enhanced" if use_ocr else "pymupdf_enhanced_chunking",
                        "extraction_method": "gpt-4-vision_ocr" if use_ocr else "pymupdf_text_extraction",
                        "tables_extracted": len(vector_store.current_tables),
                        "images_extracted": len(vector_store.current_images),
                        "attempt_number": current_attempt
                    }
                    
                    doc_ids = vector_store.add_document_with_enhanced_chunking(pdf_text, metadata)
                    if not doc_ids:
                        logger.warning(f"無法處理 {file_name}，跳過")
                        break
                    
                    if use_ocr:
                        logger.info(f"第{current_attempt}次嘗試 - OCR處理完成：提取了 {len(vector_store.current_images)} 個圖像頁面")
                    else:
                        logger.info(f"第{current_attempt}次嘗試 - 文字提取完成：{len(vector_store.current_tables)} 個表格，{len(vector_store.current_images)} 個圖像")
                    
                    success = True
                    total_processed += 1
                    logger.info(f"成功前處理 {file_name}（嘗試 {current_attempt} 次）")
                    
                except Exception as e:
                    logger.error(f"前處理 {file_name} 第 {current_attempt} 次嘗試時發生錯誤: {e}")
                    if current_attempt < max_attempts:
                        current_attempt += 1
                        continue
                    else:
                        total_failed += 1
                        break
        
        logger.info(f"{company_name} 前處理完成")
    
    logger.info(f"\n=== 前處理完成 ===")
    logger.info(f"成功處理: {total_processed} 個檔案")
    logger.info(f"處理失敗: {total_failed} 個檔案")
    
    # 檢查資料庫中的文檔數量
    total_docs = vector_store.collection.count_documents({})
    logger.info(f"向量資料庫中共有 {total_docs} 個文檔塊")
    
    return total_processed > 0

def preprocess_new_files_only(vector_store):
    """只處理新增的檔案"""
    logger.info("=== 增量處理：只處理新檔案 ===")
    
    # 獲取已處理的檔案列表
    processed_files = check_processed_files(vector_store)
    logger.info(f"已處理檔案數量: {len(processed_files)}")
    
    # 尋找所有檔案
    report_folders = find_report_folders(settings.base_directory)
    
    new_files_processed = 0
    
    for folder in report_folders:
        company_name = extract_company_name(folder)
        pdf_files = find_pdf_files(folder)
        
        for pdf_file in pdf_files:
            file_name = os.path.basename(pdf_file)
            file_key = f"{company_name}_{file_name}"
            
            # 檢查是否已經處理過
            if file_key in processed_files:
                logger.info(f"跳過已處理檔案: {file_name}")
                continue
            
            logger.info(f"處理新檔案: {file_name}")
            
            # 處理新檔案
            year, quarter = extract_year_and_quarter(file_name)
            if not year or not quarter:
                continue
            
            try:
                pdf_text, total_pages = vector_store.read_pdf_enhanced(pdf_file, company_name)
                if not pdf_text:
                    continue
                
                use_ocr = vector_store.should_use_ocr_processing(file_name, company_name)
                metadata = {
                    "file_name": file_name,
                    "company_name": company_name,
                    "year": year,
                    "quarter": f"{year}_{quarter}",
                    "total_pages": total_pages,
                    "processing_mode": "ocr_enhanced" if use_ocr else "pymupdf_enhanced_chunking",
                    "extraction_method": "gpt-4-vision_ocr" if use_ocr else "pymupdf_text_extraction",
                    "tables_extracted": len(vector_store.current_tables),
                    "images_extracted": len(vector_store.current_images),
                    "attempt_number": 1
                }
                
                doc_ids = vector_store.add_document_with_enhanced_chunking(pdf_text, metadata)
                if doc_ids:
                    new_files_processed += 1
                    logger.info(f"成功處理新檔案: {file_name}")
                
            except Exception as e:
                logger.error(f"處理新檔案 {file_name} 時發生錯誤: {e}")
    
    logger.info(f"增量處理完成，新增 {new_files_processed} 個檔案")
    return new_files_processed > 0

def analyze_companies_from_database(vector_store):
    """從資料庫中分析各公司財報"""
    logger.info("\n=== 從資料庫分析各公司財報 ===")
    
    # 創建分析資料夾
    analysis_folder = create_output_directory(settings.base_directory)
    
    # 創建Excel工作簿
    excel_filename = generate_excel_filename()
    combined_excel = os.path.join(analysis_folder, excel_filename)

    wb_combined = Workbook()
    wb_combined.remove(wb_combined.active)
    
    # 選擇是否清空舊的分析結果
    existing_analysis_count = vector_store.analysis_collection.count_documents({})
    if existing_analysis_count > 0:
        logger.info(f"發現資料庫中已有 {existing_analysis_count} 筆分析結果")
        clear_choice = input("是否要清空舊的分析結果？(y/N): ").lower()
        
        if clear_choice == 'y':
            logger.info("清空舊的分析結果...")
            vector_store.clear_analysis_collection()
        else:
            logger.info("保留現有分析結果，將使用upsert方式更新...")
    
    # 從資料庫中獲取所有公司和季度信息
    pipeline = [
        {
            "$group": {
                "_id": {
                    "company": "$metadata.company_name",
                    "quarter": "$metadata.quarter"
                }
            }
        },
        {
            "$sort": {
                "_id.company": 1,
                "_id.quarter": 1
            }
        }
    ]
    
    company_quarters = list(vector_store.collection.aggregate(pipeline))
    logger.info(f"發現 {len(company_quarters)} 個公司-季度組合")
    
    # 按公司分組
    companies_data = {}
    for item in company_quarters:
        company = item['_id']['company']
        quarter = item['_id']['quarter']
        
        if company not in companies_data:
            companies_data[company] = []
        companies_data[company].append(quarter)
    
    logger.info(f"需要分析 {len(companies_data)} 家公司")
    
    # 初始化RAG分析器
    rag_analyzer = RAGAnalyzer()
    
    # 為每家公司創建工作表並進行分析
    headers = settings.get("excel_output.headers", ["年份_季度", "公司概況", "商業策略", "風險"])
    column_widths = settings.get("excel_output.column_widths", {"A": 15, "B": 70, "C": 70, "D": 70})
    row_height = settings.get("excel_output.row_height", 200)
    
    for company_name, quarters in companies_data.items():
        logger.info(f"\n開始分析 {company_name} 的財報...")
        logger.info(f"找到季度: {quarters}")
        
        # 創建工作表
        ws = wb_combined.create_sheet(title=company_name)
        
        # 設定樣式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 設定列寬
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        row_index = 2
        
        # 分析每個季度
        for quarter in sorted(quarters):
            try:
                logger.info(f"分析 {company_name} - {quarter}")
                
                if "_" in quarter:
                    year_part, quarter_part = quarter.split("_", 1)
                    
                    # 根據季度部分判斷報告類型並創建對應的虛擬檔名
                    if quarter_part == "全年":
                        dummy_file_name = f"{company_name}_{year_part}年報.pdf"
                    elif quarter_part.startswith("Q"):
                        dummy_file_name = f"{company_name}_{year_part}{quarter_part}季報.pdf"
                    else:
                        dummy_file_name = f"{company_name}_{quarter}.pdf"
                else:
                    dummy_file_name = f"{company_name}_{quarter}.pdf"
                
                logger.info(f"使用虛擬檔名: {dummy_file_name}")
                
                # 生成分析
                analysis = rag_analyzer.generate_enhanced_business_analysis_with_fallback(
                    vector_store, 
                    dummy_file_name,  
                    company_name, 
                    quarter
                )
                
                if analysis:
                    display_quarter = quarter.replace("_", "_")
                    
                    # 填入Excel
                    ws.cell(row=row_index, column=1).value = display_quarter
                    ws.cell(row=row_index, column=2).value = analysis["company_overview"]
                    ws.cell(row=row_index, column=3).value = analysis["business_strategy"]
                    ws.cell(row=row_index, column=4).value = analysis["risks"]
                    
                    # 設定自動換行和行高
                    for col in range(1, 5):
                        cell = ws.cell(row=row_index, column=col)
                        cell.alignment = Alignment(wrapText=True, vertical='top')
                    
                    # 設定行高
                    ws.row_dimensions[row_index].height = row_height
                    
                    # 保存分析結果到MongoDB
                    saved_docs = vector_store.save_analysis_to_mongodb(
                        company=company_name,
                        quarter=quarter,
                        analysis_results=analysis
                    )
                    
                    logger.info(f"成功保存 {len(saved_docs)} 個分析結果到MongoDB")
                    for doc in saved_docs:
                        logger.info(f"  - {doc['action']}: {doc['company']} - {doc['title']} - {doc['quarter']}")
                    
                    row_index += 1
                    logger.info(f"完成 {company_name} - {quarter} 的分析")
                    
                else:
                    logger.warning(f"跳過 {company_name} - {quarter} - 無法生成有效分析")
                
            except Exception as e:
                logger.error(f"分析 {company_name} - {quarter} 時發生錯誤: {e}")
                continue
        
        logger.info(f"{company_name} 財報分析完成")
    
    # 儲存Excel
    try:
        wb_combined.save(combined_excel)
        logger.info(f"\n財報分析完成！")
        logger.info(f"Excel結果已保存到: {combined_excel}")
        
        # 顯示MongoDB保存的統計信息
        total_analysis_count = vector_store.analysis_collection.count_documents({})
        logger.info(f"MongoDB中共保存了 {total_analysis_count} 筆分析結果")
        
        # 顯示OCR降級統計
        if vector_store.netmarble_failed_files:
            logger.info(f"使用OCR降級處理的檔案數量: {len(vector_store.netmarble_failed_files)}")
            for failed_file in vector_store.netmarble_failed_files:
                logger.info(f"  - {failed_file}")
    
    except Exception as e:
        logger.error(f"保存 Excel 時發生錯誤: {e}")

def main():
    """主程式"""
    logger.info("=== 財報分析系統啟動 ===")
    
    # 驗證設定
    if not settings.validate_required_settings():
        logger.error("設定驗證失敗，程式結束")
        return
    
    # 測試連接
    if not test_connections():
        logger.error("連接測試失敗，程式結束")
        return
    
    # 初始化向量資料庫
    vector_store = EnhancedMongoDBVectorStore()
    
    # 檢查現有資料
    existing_docs = vector_store.collection.count_documents({})
    existing_analysis = vector_store.analysis_collection.count_documents({})
    
    logger.info(f"目前資料庫狀態:")
    logger.info(f"- 向量文檔: {existing_docs} 個")
    logger.info(f"- 分析結果: {existing_analysis} 個")
    
    # 選擇處理模式
    print("\n請選擇處理模式:")
    print("1. 完整重新處理 (清空所有資料重新開始)")
    print("2. 增量處理 (只處理新檔案)")
    print("3. 只重新分析 (使用現有向量資料)")
    print("4. 退出")
    
    choice = input("請輸入選項 (1-4): ").strip()
    
    if choice == "1":
        # 完整重新處理
        success = preprocess_all_companies(vector_store, force_reprocess=True)
        if success:
            analyze_companies_from_database(vector_store)
    
    elif choice == "2":
        # 增量處理
        success = preprocess_new_files_only(vector_store)
        if success or existing_docs > 0:
            analyze_companies_from_database(vector_store)
    
    elif choice == "3":
        # 只重新分析
        if existing_docs > 0:
            analyze_companies_from_database(vector_store)
        else:
            logger.error("沒有找到現有的向量資料，請先進行檔案處理")
    
    elif choice == "4":
        logger.info("程式結束")
        return
    
    else:
        logger.error("無效選項，程式結束")
        return

if __name__ == "__main__":
    main()